#### This tool is a work in progress. It takes an excel file 'a' and compares the data against 2 reference excel files 'b' amd 'c'. File a should have data in column a , files b anad c can have data in columns a and b ####


# Set up
import openpyxl 


# Open excel docs with openpyxl, load each sheet.
###TO DO - Transfer each workbook to sheets in 1 book, open just the 1 workbook, give better names###
species=openpyxl.load_workbook("species.xlsx")
species_sheet = species['Sheet1']
ref = openpyxl.load_workbook("Itis_ref.xlsx")
ref_sheet = ref['Sheet1']
iti_ref = openpyxl.load_workbook("ITI_ref.xlsx")
iti_ref_sheet = iti_ref['Sheet1']
iti_list = openpyxl.load_workbook("iti_list.xlsx")
iti_list_sheet = iti_list['Sheet1']

# ITIS scores
# Construct dictionary from file b, with column a as key and column b as value

ITIS_dict = {}

for row in range(2, ref_sheet.max_row + 1):
    species_name = ref_sheet['A' + str(row)].value
    ITI_score = ref_sheet['B' + str(row)].value
    ITIS_dict.setdefault(species_name,ITI_score)

print(ITIS_dict)

#iterate over file a, if data is found in file b, copy the data from file b column b into file a column b.

for rowNum in range(1, species_sheet.max_row+1):
    species_name_org = species_sheet.cell(row=rowNum, column=1).value
    if species_name_org in ITIS_dict:
        x = ITIS_dict.get(species_name_org)
        print(x)
        species_sheet.cell(row=rowNum, column=2).value = x


# ITI Name
# Construct dictionary from file c, with column a as key and column b as value

ITI_dict = {}

for row in range(2, iti_ref_sheet.max_row + 1):
    species_name = iti_ref_sheet['A' + str(row)].value
    ITI_name = iti_ref_sheet['B' + str(row)].value
    ITI_dict.setdefault(species_name,ITI_name)

print(ITI_dict)

    
#iterate over file a, if data is found in file c, copy the data from file b column b into file a column c.

for rowNum in range(1, species_sheet.max_row+1):
    species_name_org = species_sheet.cell(row=rowNum, column=1).value
    if species_name_org in ITI_dict.keys():
        x = ITI_dict.get(species_name_org)
        print(x)
        species_sheet.cell(row=rowNum, column=3).value = x
    else:
        for row in iti_list_sheet.iter_rows(values_only=True):
            for rowNum in range(1, species_sheet.max_row+1):
                species_name = species_sheet.cell(row=rowNum, column=1).value
                if species_name in row:
                    species_sheet.cell(row=rowNum, column=3).value = species_sheet.cell(row=rowNum, column=1).value

species.save("output.xlsx")
